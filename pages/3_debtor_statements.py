import streamlit as st
import re
import io
import zipfile
import pandas as pd
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- 1. SECURITY BOUNCER ---
if "password_correct" not in st.session_state or not st.session_state["password_correct"]:
    st.warning("🔒 Connection lost or not logged in.")
    st.info("Please click the Main Portal page in your sidebar to log in and reconnect to the database.")
    st.stop()

st.title("🧾 Debtor Statement Generator")
st.write(
    "Upload your full Debtors Ledger export. This tool will clean it up, drop any "
    "debtor with a nil balance, let you filter and preview the remaining debtors, "
    "and produce formatted statements — split into one zip per balance range you define."
)

# --- 2. FORMATTING CONSTANTS (matches the updated R R Metal reference format) ---
ACCT_FMT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
DATE_FMT = 'mm-dd-yy'
HEADER_FILL = 'FFC0C0C0'
THIN = Side(style='thin')
FULL_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def clean_name(raw):
    """Strip the leading 'CODE - ' prefix from a debtor name."""
    raw = str(raw).strip()
    if ' - ' in raw:
        return raw.split(' - ', 1)[1].strip()
    return raw


def get_code(raw):
    """Extract the leading debtor code (before ' - ')."""
    raw = str(raw).strip()
    if ' - ' in raw:
        return raw.split(' - ', 1)[0].strip()
    return ''


def sanitize_filename(name):
    """Make a debtor name safe to use as a filename."""
    name = re.sub(r'[\\/*?:"<>|]', '', name).strip()
    name = re.sub(r'\s+', ' ', name)
    return name[:150] if name else "Unnamed"


def extract_ledger_header(ws):
    """Looks for the 6-row company/report header block that Tally sometimes
    puts above the actual 'Date | Particulars | Debit | Credit | Balance'
    column-header row (company name, address, country, a blank spacer, the
    report title, and the period caption - matching the R R Metal reference
    file's rows 1-6). Returns (header_values, data_start_row):
      - header_values is None if the uploaded ledger starts directly with
        the column-header row (no metadata block present) - fully
        backward compatible with ledgers that don't have this block.
      - data_start_row is the row where 'Date'/'Particulars' actually is,
        so the rest of the parser knows where the real data begins
        regardless of whether a header block was found.
    """
    # Find the real column-header row - don't assume it's row 1.
    header_row_idx = None
    for r in range(1, min(ws.max_row, 20) + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if str(a).strip() == 'Date' and str(b).strip() == 'Particulars':
            header_row_idx = r
            break

    if header_row_idx is None:
        # Couldn't find it at all in the first 20 rows - treat the whole
        # sheet as headerless data starting at row 1, same as before.
        return None, 1

    if header_row_idx == 1:
        # No metadata block - ledger starts directly with column headers.
        return None, 1

    if header_row_idx == 7:
        # Exactly matches the known 6-row Tally header block layout.
        header_values = {
            'company_name': ws.cell(1, 1).value,
            'fy_label': ws.cell(1, 3).value,
            'address': ws.cell(2, 1).value,
            'date_range': ws.cell(2, 3).value,
            'country': ws.cell(3, 1).value,
            'report_title': ws.cell(5, 1).value,
            'period_caption': ws.cell(6, 1).value,
        }
        return header_values, header_row_idx

    # Some other offset we don't recognize - safer to skip the block
    # entirely than guess at a layout that might not match.
    return None, header_row_idx


def parse_debtors(file_obj):
    """Parse the Tally-style debtors ledger into a list of per-debtor blocks."""
    wb = load_workbook(file_obj, data_only=False)
    ws = wb[wb.sheetnames[0]]

    ledger_header, data_start_row = extract_ledger_header(ws)

    blocks = []
    current = None

    for r in range(data_start_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        e = ws.cell(r, 5).value

        if a is not None and str(a).strip() == 'Customer:':
            if current:
                blocks.append(current)
            current = {
                'raw_name': str(b).strip() if b is not None else 'Unnamed',
                'opening': None,
                'txns': [],
                'final_balance_text': None,
            }
            continue

        if current is None:
            continue

        b_str = str(b).strip() if b is not None else ''

        if b_str == 'Opening Balance...':
            current['opening'] = c if c is not None else 0
            continue

        if b_str == 'Party Total =>>':
            current['final_balance_text'] = str(e) if e is not None else ''
            continue

        if a is not None or b is not None:
            current['txns'].append({
                'date': a, 'particulars': b, 'debit': c, 'credit': d, 'bal_text': e
            })

    if current:
        blocks.append(current)

    return blocks, ledger_header


def is_nil(block):
    """A debtor is nil if their final recorded balance says 'Nil'."""
    text = block['final_balance_text']
    if text is None and block['txns']:
        text = block['txns'][-1]['bal_text']
    if text is None:
        return True
    return 'nil' in str(text).lower()


def get_final_balance(block):
    """Signed numeric balance: positive = Dr (owes us), negative = Cr (we owe them)."""
    text = block['final_balance_text']
    if text is None and block['txns']:
        text = block['txns'][-1]['bal_text']
    if text is None:
        return 0.0
    text = str(text).strip()
    num_match = re.search(r'[\d,]+\.?\d*', text)
    if not num_match:
        return 0.0
    val = float(num_match.group().replace(',', ''))
    return -val if 'cr' in text.lower() else val


def get_last_txn_date(block):
    """Most recent transaction date for this debtor, or None if no dated transactions."""
    dates = [t['date'] for t in block['txns'] if isinstance(t['date'], datetime)]
    return max(dates) if dates else None


def style_cell(cell, bold=False, color=None, align=None, numfmt=None, fill=None, border=True):
    cell.font = Font(name='Arial', size=10, bold=bold, color=color)
    if align:
        cell.alignment = Alignment(horizontal=align)
    if numfmt:
        cell.number_format = numfmt
    if fill:
        cell.fill = PatternFill('solid', start_color=fill, end_color=fill)
    if border:
        cell.border = FULL_BORDER


def build_workbook(block, ledger_header=None):
    """Build a single-debtor statement workbook, formatted to match the updated
    R R Metal reference file: an optional 6-row company/report header block
    (extracted from the uploaded ledger, if present), then the column-header
    row, then the debtor's own title row, full thin borders throughout."""
    title = clean_name(block['raw_name'])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    ws.column_dimensions['A'].width = 10.14
    ws.column_dimensions['B'].width = 23.14
    ws.column_dimensions['C'].width = 12.86
    ws.column_dimensions['D'].width = 11.29
    ws.column_dimensions['E'].width = 12.86

    offset = 0
    if ledger_header:
        # Rows 1-6: company name/FY, address/date-range, country, a blank
        # spacer, the report title, and the period caption. Matches the R R
        # Metal reference file's rows 1-6 exactly - including that, unlike
        # the rest of the sheet, these rows have NO border and NO fill;
        # only the labeled cells (A and C) carry real font styling, the
        # rest of each merged region is just plain default Calibri 11.
        def _plain(cell):
            cell.font = Font(name='Calibri', size=11)

        ws.merge_cells('A1:B1')
        ws['A1'] = ledger_header.get('company_name')
        ws['A1'].font = Font(name='Arial', size=12, bold=True)
        _plain(ws['B1'])

        ws.merge_cells('C1:E1')
        ws['C1'] = ledger_header.get('fy_label')
        ws['C1'].font = Font(name='Arial', size=12, bold=True)
        ws['C1'].alignment = Alignment(horizontal='right')
        ws['C1'].number_format = ACCT_FMT
        _plain(ws['D1']); _plain(ws['E1'])

        ws.merge_cells('A2:B2')
        ws['A2'] = ledger_header.get('address')
        ws['A2'].font = Font(name='Calibri', size=10)
        _plain(ws['B2'])

        ws.merge_cells('C2:E2')
        ws['C2'] = ledger_header.get('date_range')
        ws['C2'].font = Font(name='Calibri', size=10)
        ws['C2'].alignment = Alignment(horizontal='right')
        ws['C2'].number_format = ACCT_FMT
        _plain(ws['D2']); _plain(ws['E2'])

        ws.merge_cells('A3:B3')
        ws['A3'] = ledger_header.get('country')
        ws['A3'].font = Font(name='Calibri', size=10)
        _plain(ws['B3'])

        ws.merge_cells('C3:D3')
        ws['C3'].font = Font(name='Calibri', size=10)
        ws['C3'].alignment = Alignment(horizontal='right')
        ws['C3'].number_format = ACCT_FMT
        _plain(ws['E3'])

        ws.merge_cells('A4:E4')
        ws['A4'].font = Font(name='Calibri', size=7.5, bold=True)
        ws['A4'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A5:E5')
        ws['A5'] = ledger_header.get('report_title')
        ws['A5'].font = Font(name='Calibri', size=12, bold=True)
        ws['A5'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A6:E6')
        ws['A6'] = ledger_header.get('period_caption')
        ws['A6'].font = Font(name='Calibri', size=10)
        ws['A6'].alignment = Alignment(horizontal='center')

        offset = 6

    # Column headers
    header_row = offset + 1
    headers = ['Date', 'Particulars', 'Debit', 'Credit', 'Balance']
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, text)
        align = 'right' if col in (3, 4, 5) else None
        style_cell(cell, bold=True, color='FF000000', align=align, fill=HEADER_FILL,
                   numfmt=ACCT_FMT if col in (3, 4, 5) else None)

    # Title row + running-balance seed
    title_row = offset + 2
    ws.merge_cells(f'A{title_row}:D{title_row}')
    ws[f'A{title_row}'] = title
    style_cell(ws[f'A{title_row}'], bold=True, color='FF0000FF', align='center', border=False, fill='FFFFFFFF')
    ws[f'A{title_row}'].border = Border(top=THIN, bottom=THIN, left=THIN, right=None)
    for col_letter in ('B', 'C'):
        coord = f'{col_letter}{title_row}'
        ws[coord].border = Border(top=THIN, bottom=THIN, left=None, right=None)
        ws[coord].font = Font(name='Arial', size=10)
        ws[coord].fill = PatternFill('solid', start_color='FFFFFFFF', end_color='FFFFFFFF')
    ws[f'D{title_row}'].border = Border(top=THIN, bottom=THIN, left=None, right=THIN)
    ws[f'D{title_row}'].font = Font(name='Arial', size=10)
    ws[f'D{title_row}'].fill = PatternFill('solid', start_color='FFFFFFFF', end_color='FFFFFFFF')

    ws[f'E{title_row}'] = 0
    style_cell(ws[f'E{title_row}'], align='right', numfmt=ACCT_FMT, fill='FFFFFFFF')

    row = title_row + 1
    if block['opening'] is not None:
        ws.cell(row, 2, 'Opening Balance...')
        ws.cell(row, 3, block['opening'])
        ws.cell(row, 5, f'=E{row-1}+C{row}-D{row}')
        style_cell(ws.cell(row, 1), fill='FFFFFFFF')
        style_cell(ws.cell(row, 2), fill='FFFFFFFF')
        style_cell(ws.cell(row, 3), align='right', numfmt=ACCT_FMT, fill='FFFFFFFF')
        style_cell(ws.cell(row, 4), align='right', numfmt=ACCT_FMT, fill='FFFFFFFF')
        style_cell(ws.cell(row, 5), align='right', numfmt=ACCT_FMT, fill='FFFFFFFF')
        row += 1

    for txn in block['txns']:
        ws.cell(row, 1, txn['date'])
        style_cell(ws.cell(row, 1), numfmt=DATE_FMT, fill='FFFFFFFF')

        ws.cell(row, 2, txn['particulars'])
        style_cell(ws.cell(row, 2), fill='FFFFFFFF')

        if txn['debit']:
            ws.cell(row, 3, txn['debit'])
        style_cell(ws.cell(row, 3), align='right', numfmt=ACCT_FMT, fill='FFFFFFFF')

        if txn['credit']:
            ws.cell(row, 4, txn['credit'])
        style_cell(ws.cell(row, 4), align='right', numfmt=ACCT_FMT, fill='FFFFFFFF')

        ws.cell(row, 5, f'=E{row-1}+C{row}-D{row}')
        style_cell(ws.cell(row, 5), align='right', numfmt=ACCT_FMT, fill='FFFFFFFF')
        row += 1

    last_row = row - 1
    if last_row >= title_row + 1:
        style_cell(ws.cell(last_row, 5), bold=True, align='right', numfmt=ACCT_FMT, fill='FFFFFFFF')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, title


def build_summary_workbook(rows_df):
    """Master summary sheet listing every included debtor."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'

    widths = {'A': 12, 'B': 45, 'C': 16, 'D': 14, 'E': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    headers = ['Code', 'Debtor Name', 'Final Balance', 'Last Transaction', 'Days Since']
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(1, col, text)
        align = 'right' if col in (3, 5) else None
        style_cell(cell, bold=True, align=align, fill=HEADER_FILL)

    row = 2
    for _, r in rows_df.iterrows():
        ws.cell(row, 1, r['Code'])
        style_cell(ws.cell(row, 1))
        ws.cell(row, 2, r['Debtor Name'])
        style_cell(ws.cell(row, 2))
        ws.cell(row, 3, r['Final Balance'])
        style_cell(ws.cell(row, 3), align='right', numfmt=ACCT_FMT)
        if pd.notna(r['Last Transaction']):
            ws.cell(row, 4, r['Last Transaction'])
            style_cell(ws.cell(row, 4), numfmt=DATE_FMT)
        else:
            style_cell(ws.cell(row, 4))
        ws.cell(row, 5, r['Days Since'] if pd.notna(r['Days Since']) else None)
        style_cell(ws.cell(row, 5), align='right')
        row += 1

    total_row = row
    ws.cell(total_row, 2, 'TOTAL')
    style_cell(ws.cell(total_row, 2), bold=True)
    ws.cell(total_row, 3, f'=SUM(C2:C{row-1})')
    style_cell(ws.cell(total_row, 3), bold=True, align='right', numfmt=ACCT_FMT)
    ws.cell(total_row, 1, f'=COUNTA(A2:A{row-1})')
    style_cell(ws.cell(total_row, 1), bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --- 3. UI: UPLOAD & PARSE ---
uploaded_file = st.file_uploader("Upload Debtors Ledger (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    if "debtor_blocks" not in st.session_state or st.session_state.get("debtor_file_name") != uploaded_file.name:
        with st.spinner("Reading and parsing the ledger..."):
            try:
                st.session_state.debtor_blocks, st.session_state.ledger_header = parse_debtors(uploaded_file)
                st.session_state.debtor_file_name = uploaded_file.name
            except Exception as e:
                st.error(f"Failed to read the file: {e}")
                st.stop()

    blocks = st.session_state.debtor_blocks
    ledger_header = st.session_state.get("ledger_header")
    total = len(blocks)
    non_nil_blocks = [b for b in blocks if not is_nil(b)]
    nil_count = total - len(non_nil_blocks)

    if ledger_header:
        st.caption(
            f"📋 Detected a company/report header block in the uploaded ledger "
            f"({ledger_header.get('company_name', '').strip()}) — this will be "
            f"reproduced at the top of every generated statement."
        )

    st.success(
        f"✅ Parsed {total} debtors — {nil_count} had a nil balance and were removed, "
        f"{len(non_nil_blocks)} remain."
    )

    if not non_nil_blocks:
        st.warning("No debtors with an outstanding balance were found.")
        st.stop()

    # --- 4. CURRENT DATE INPUT (for Days Since calculation) ---
    st.header("📅 Set Current Date")
    st.caption(
        "Enter the date to calculate 'Days Since Last Purchase' against. Since your ledger "
        "dates are Bikram Sambat numbers (not real Gregorian dates), enter the equivalent "
        "BS date here — the tool does simple day-count arithmetic against it, not a real "
        "calendar conversion."
    )
    current_date = st.date_input(
        "Current date",
        value=date.today(),
        min_value=date(2000, 1, 1),
        max_value=date(2150, 12, 31),
    )

    # --- 5. BUILD METRICS TABLE ---
    records = []
    for b in non_nil_blocks:
        last_date = get_last_txn_date(b)
        days_since = (current_date - last_date.date()).days if last_date else None
        records.append({
            'Code': get_code(b['raw_name']),
            'Debtor Name': clean_name(b['raw_name']),
            'Final Balance': get_final_balance(b),
            'Last Transaction': last_date,
            'Days Since': days_since,
            '_block': b,
        })
    df_all = pd.DataFrame(records)

    # --- 6. SEARCH ---
    st.header("🔍 Search")
    search_term = st.text_input("Search debtor name", placeholder="Type to filter by name...")

    df_search = df_all
    if search_term:
        df_search = df_search[df_search['Debtor Name'].str.contains(search_term, case=False, na=False)]

    bal_min, bal_max = float(df_search['Final Balance'].min()), float(df_search['Final Balance'].max())
    st.caption(
        f"Debit (Dr.) balances are positive, Credit (Cr.) balances are negative. "
        f"Balances currently range from {bal_min:,.2f} to {bal_max:,.2f}."
    )

    # --- 7. MULTIPLE BALANCE RANGES ---
    st.header("🎚️ Define Balance Ranges")
    st.caption(
        "Each range you define below will be generated as its own zip file "
        "(with its own summary sheet and individual statements)."
    )

    if "num_ranges" not in st.session_state:
        st.session_state.num_ranges = 1

    rc1, rc2, rc3 = st.columns([1, 1, 2])
    if rc1.button("➕ Add Range"):
        st.session_state.num_ranges += 1
    if rc2.button("➖ Remove Range") and st.session_state.num_ranges > 1:
        st.session_state.num_ranges -= 1

    ranges = []
    for i in range(st.session_state.num_ranges):
        col1, col2 = st.columns(2)
        r_min = col1.number_input(
            f"Range {i + 1} — Minimum Balance", value=bal_min, key=f"range_min_{i}"
        )
        r_max = col2.number_input(
            f"Range {i + 1} — Maximum Balance", value=bal_max, key=f"range_max_{i}"
        )
        ranges.append((r_min, r_max))

    # --- 8. PREVIEW PER RANGE ---
    st.header("👀 Preview")
    range_dfs = []
    for i, (r_min, r_max) in enumerate(ranges):
        df_range = df_search[
            (df_search['Final Balance'] >= r_min) & (df_search['Final Balance'] <= r_max)
        ]
        range_dfs.append(df_range)

        with st.expander(
            f"Range {i + 1}: {r_min:,.2f} to {r_max:,.2f} — "
            f"{len(df_range)} debtors, total {df_range['Final Balance'].sum():,.2f}"
        ):
            st.dataframe(
                df_range.drop(columns=['_block']).style.format({
                    'Final Balance': '{:,.2f}',
                    'Last Transaction': lambda d: d.strftime('%Y-%m-%d') if pd.notna(d) else '',
                }),
                use_container_width=True,
                hide_index=True,
            )

    total_selected = sum(len(d) for d in range_dfs)
    if total_selected == 0:
        st.warning("No debtors fall within the defined ranges.")
        st.stop()

    # --- 9. GENERATE ---
    if st.button("⚙️ Generate Zip Files"):
        with st.spinner("Generating statements for each range..."):
            master_zip_buf = io.BytesIO()
            with zipfile.ZipFile(master_zip_buf, "w", zipfile.ZIP_DEFLATED) as master_zf:
                progress = st.progress(0)
                done = 0
                for i, df_range in enumerate(range_dfs):
                    r_min, r_max = ranges[i]
                    folder = f"Range_{i + 1} ({r_min:,.2f} to {r_max:,.2f})"

                    if df_range.empty:
                        continue

                    summary_buf = build_summary_workbook(df_range)
                    master_zf.writestr(f"{folder}/0_Summary.xlsx", summary_buf.getvalue())

                    used_names = {}
                    for _, r in df_range.iterrows():
                        file_buf, title = build_workbook(r['_block'], ledger_header)
                        fname = sanitize_filename(title)
                        if fname in used_names:
                            used_names[fname] += 1
                            fname = f"{fname} ({used_names[fname]})"
                        else:
                            used_names[fname] = 0
                        master_zf.writestr(f"{folder}/{fname}.xlsx", file_buf.getvalue())

                        done += 1
                        progress.progress(done / total_selected)
            master_zip_buf.seek(0)

        st.success(f"🎉 Generated {len(ranges)} range folder(s) inside one zip!")
        st.download_button(
            label="⬇️ Download All Ranges (.zip)",
            data=master_zip_buf,
            file_name="Debtor_Statements_By_Range.zip",
            mime="application/zip",
        )
else:
    st.info("Upload your Debtors Ledger export to get started.")


# --- 10. BALANCE CONFIRMATION PDF SPLITTER ---
st.divider()
st.header("📄 Balance Confirmation PDF Splitter")
st.write(
    "Upload a consolidated Balance Confirmation PDF (one confirmation letter + "
    "ledger per debtor, all in a single file). This will drop any debtor with "
    "**both** a zero closing balance **and** a fiscal-year turnover under "
    "Rs 100,000, then split the rest into one PDF per debtor — keeping the "
    "exact same letter + ledger format as the original."
)

bc_pdf = st.file_uploader("Upload Balance Confirmation PDF", type=["pdf"], key="bc_pdf_uploader")

TURNOVER_THRESHOLD = 100000


def parse_balance_confirmation_pdf(file_obj):
    """Splits the consolidated PDF into per-debtor page ranges and pulls out
    each debtor's name, turnover ('Total Transaction RS'), and closing
    balance ('Current Balance till date') from their letter page.

    Uses the pdftotext CLI rather than a Python PDF library for the text
    extraction step - on a large consolidated file (this kind of export can
    run to 1000+ pages) a pure-Python page-by-page extraction library holds
    far more in memory than this environment has available and gets killed;
    pdftotext streams through the file instead."""
    import subprocess
    import tempfile
    import os

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(file_obj.read())
        tmp_path = tmp.name

    try:
        result = subprocess.run(
            ["pdftotext", "-layout", tmp_path, "-"],
            capture_output=True, text=True, check=True,
        )
        full_text = result.stdout
    finally:
        os.unlink(tmp_path)

    pages_text = full_text.split("\x0c")
    if pages_text and pages_text[-1].strip() == "":
        pages_text = pages_text[:-1]

    debtor_starts = []
    for i, p in enumerate(pages_text):
        m = re.search(r'^M/s (.+)$', p, re.MULTILINE)
        if m:
            debtor_starts.append((i, m.group(1).strip()))

    records = []
    for idx, (page_i, name) in enumerate(debtor_starts):
        page_text = pages_text[page_i]

        turnover_m = re.search(r'Total Transaction RS:\s*([\d,]+\.?\d*)', page_text)
        turnover = float(turnover_m.group(1).replace(',', '')) if turnover_m else 0.0

        bal_m = re.search(r'Current Balance till date:\s*RS\s*([\d,]+\.?\d*)\s*(Dr|Cr)', page_text)
        if bal_m:
            val = float(bal_m.group(1).replace(',', ''))
            closing_balance = -val if bal_m.group(2) == 'Cr' else val
        else:
            closing_balance = None

        name_match = re.search(r'^M/s .+$', page_text, re.MULTILINE)
        text_after_name = page_text[name_match.end():] if name_match else page_text
        pan_m = re.search(r'PAN No\.:\s*(\S*)', text_after_name)
        pan = pan_m.group(1) if pan_m else ''

        end_page = debtor_starts[idx + 1][0] - 1 if idx + 1 < len(debtor_starts) else len(pages_text) - 1

        records.append({
            'name': name,
            'pan': pan,
            'start_page': page_i,
            'end_page': end_page,
            'turnover': turnover,
            'closing_balance': closing_balance,
        })

    return records


def get_highlighted_names(file_obj):
    """Reads an Excel file and returns the set of (uppercased, trimmed) text
    values found in any cell that has a solid, non-white/non-default fill -
    i.e. manually highlighted in Excel. Checks every sheet, since the user
    might not have the debtor names on the first one."""
    wb = load_workbook(file_obj, data_only=True)
    names = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                fill = cell.fill
                if fill is None or fill.patternType != 'solid':
                    continue
                fg = fill.fgColor
                rgb = fg.rgb if fg else None
                is_theme_colored = bool(fg) and getattr(fg, 'type', None) == 'theme'
                is_rgb_colored = rgb not in (None, '00000000', 'FFFFFFFF')
                if is_rgb_colored or is_theme_colored:
                    names.add(cell.value.strip().upper())
    return names


def build_bc_summary_workbook(records):
    """Excel summary of the selected (kept) debtors from the Balance
    Confirmation PDF, including their final (closing) balance."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'

    widths = {'A': 45, 'B': 16, 'C': 16, 'D': 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    headers = ['Debtor Name', 'PAN', 'Turnover', 'Final Balance']
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(1, col, text)
        align = 'right' if col in (3, 4) else None
        style_cell(cell, bold=True, align=align, fill=HEADER_FILL)

    row = 2
    for r in records:
        ws.cell(row, 1, r['name'])
        style_cell(ws.cell(row, 1))
        ws.cell(row, 2, r['pan'])
        style_cell(ws.cell(row, 2))
        ws.cell(row, 3, r['turnover'])
        style_cell(ws.cell(row, 3), align='right', numfmt=ACCT_FMT)
        ws.cell(row, 4, r['closing_balance'])
        style_cell(ws.cell(row, 4), align='right', numfmt=ACCT_FMT)
        row += 1

    total_row = row
    ws.cell(total_row, 1, 'TOTAL')
    style_cell(ws.cell(total_row, 1), bold=True)
    ws.cell(total_row, 3, f'=SUM(C2:C{row-1})')
    style_cell(ws.cell(total_row, 3), bold=True, align='right', numfmt=ACCT_FMT)
    ws.cell(total_row, 4, f'=SUM(D2:D{row-1})')
    style_cell(ws.cell(total_row, 4), bold=True, align='right', numfmt=ACCT_FMT)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


if bc_pdf is not None:
    if ("bc_records" not in st.session_state
            or st.session_state.get("bc_file_name") != bc_pdf.name):
        with st.spinner("Reading the PDF and locating each debtor's block..."):
            try:
                st.session_state.bc_records = parse_balance_confirmation_pdf(bc_pdf)
                st.session_state.bc_file_name = bc_pdf.name
                bc_pdf.seek(0)
                st.session_state.bc_pdf_bytes = bc_pdf.read()
            except Exception as e:
                st.error(f"Failed to read the PDF: {e}")
                st.stop()

    bc_records = st.session_state.bc_records
    st.success(f"✅ Found {len(bc_records)} debtors in the uploaded PDF.")

    missing_balance = [r for r in bc_records if r['closing_balance'] is None]
    if missing_balance:
        st.warning(
            f"⚠️ Couldn't find a closing balance for {len(missing_balance)} debtor(s) — "
            f"they'll be kept by default rather than risk dropping them incorrectly: "
            + ", ".join(r['name'] for r in missing_balance[:10])
            + ("..." if len(missing_balance) > 10 else "")
        )

    bc_auto_excluded = [
        r for r in bc_records
        if r['closing_balance'] == 0 and r['turnover'] < TURNOVER_THRESHOLD
    ]
    bc_after_auto_filter = [r for r in bc_records if r not in bc_auto_excluded]

    st.subheader("🚫 Optional: Skip Specific Debtors")
    st.caption(
        "Upload an Excel file with some debtor names highlighted (any solid "
        "fill color, in any sheet/column). Those debtors will be skipped "
        "from the final list and split PDFs, on top of the balance/turnover "
        "filter above."
    )
    exclusion_file = st.file_uploader(
        "Upload Excel with highlighted debtors to skip (optional)",
        type=["xlsx"],
        key="bc_exclusion_uploader",
    )

    bc_manually_excluded = []
    bc_kept = bc_after_auto_filter
    if exclusion_file is not None:
        try:
            highlighted_names = get_highlighted_names(exclusion_file)
            bc_manually_excluded = [
                r for r in bc_after_auto_filter
                if r['name'].strip().upper() in highlighted_names
            ]
            bc_kept = [r for r in bc_after_auto_filter if r not in bc_manually_excluded]
            if highlighted_names and not bc_manually_excluded:
                st.warning(
                    "Found highlighted cells, but none of the names matched a debtor "
                    "in this PDF — check that the highlighted text matches the debtor "
                    "names exactly."
                )
        except Exception as e:
            st.error(f"Failed to read the exclusion file: {e}")

    c1, c2, c3 = st.columns(3)
    c1.metric("Debtors Kept", len(bc_kept))
    c2.metric("Removed (zero balance + turnover < 100,000)", len(bc_auto_excluded))
    c3.metric("Removed (manually highlighted)", len(bc_manually_excluded))

    bc_df = pd.DataFrame([
        {
            'Name': r['name'],
            'PAN': r['pan'],
            'Turnover': r['turnover'],
            'Closing Balance': r['closing_balance'],
            'Pages': r['end_page'] - r['start_page'] + 1,
        }
        for r in bc_kept
    ])
    st.dataframe(
        bc_df.style.format({'Turnover': '{:,.2f}', 'Closing Balance': '{:,.2f}'}),
        use_container_width=True,
        hide_index=True,
    )

    if st.button("⚙️ Split PDF Into Per-Debtor Files", key="bc_split_button"):
        with st.spinner(f"Splitting into {len(bc_kept)} individual PDFs..."):
            from pypdf import PdfReader, PdfWriter

            reader = PdfReader(io.BytesIO(st.session_state.bc_pdf_bytes))

            zip_buf = io.BytesIO()
            used_names = {}
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                summary_buf = build_bc_summary_workbook(bc_kept)
                zf.writestr("0_Summary.xlsx", summary_buf.getvalue())

                progress = st.progress(0)
                for i, r in enumerate(bc_kept):
                    writer = PdfWriter()
                    for p in range(r['start_page'], r['end_page'] + 1):
                        writer.add_page(reader.pages[p])

                    out_buf = io.BytesIO()
                    writer.write(out_buf)
                    out_buf.seek(0)

                    fname = sanitize_filename(r['name'])
                    if fname in used_names:
                        used_names[fname] += 1
                        fname = f"{fname} ({used_names[fname]})"
                    else:
                        used_names[fname] = 0

                    zf.writestr(f"{fname}.pdf", out_buf.getvalue())
                    progress.progress((i + 1) / len(bc_kept))
            zip_buf.seek(0)

        st.success("🎉 All balance confirmation PDFs and the summary sheet are ready!")
        st.download_button(
            label="⬇️ Download Per-Debtor Confirmations + Summary (.zip)",
            data=zip_buf,
            file_name="Balance_Confirmations_By_Debtor.zip",
            mime="application/zip",
            key="bc_download_button",
        )
else:
    st.info("Upload a Balance Confirmation PDF to get started.")
